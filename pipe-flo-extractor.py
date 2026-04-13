"""
=============================================================================
PipeFlo Extractor  —  ESI PipeFlo v17.1
=============================================================================
Extrae de archivos .pipe:
  · Cañerías  : nombre, D.nominal, OD, WT, ID, longitud, material, rugosidad,
                nodo inicio/fin, fittings (Le/D o K), K_total.
  · Nodos     : nombre, elevación (m), posición en grilla (X, Y).
  · Fluido    : nombre, temperatura, modelo, propiedades físicas.
  · Componentes especiales: Tank, Pump, Fixed dP, Control Valve, Pressure Boundary.

K_TOTAL:
  Los valores K calculados por el solver PipeFlo se ingresan en USER_K_OVERRIDES.
  Si no hay override, se calcula con f_T × Le/D (Crane TP-410, zona turbulenta).

Función principal de integración:
  export_variables(filepath)  →  dict con TODOS los datos listos para usar
                                  en cualquier programa externo.

Uso interactivo: python pipeflo_extractor.py
=============================================================================
"""

import re, os, sys, csv, json, math

# ─────────────────────────────────────────────────────────────────────────────
# K de singularidades  —  valores del solver PipeFlo (provistos por usuario)
# ─────────────────────────────────────────────────────────────────────────────
USER_K_OVERRIDES: dict = {
    # Valores K calculados por PipeFlo (del solver hidráulico a caudal real).
    # Agrega o modifica según los resultados que PipeFlo reporte.
    'Pipe 1' : 1.028,   # Ball + Reducer contraction 100 mm
    'Pipe 3' : 2.278,   # 2×Mitre 90° + 2×Mitre 45°
    'Pipe 4' : 1.822,   # 2×Mitre 90°
    'Pipe 13': 0.0,
    'Pipe 17': 0.0,
    'Pipe 28': 3.286,   # Ball + Butterfly + 2×Reducer 110 mm
}

# ─────────────────────────────────────────────────────────────────────────────
# Tabla de materiales de cañería (spec_obj_id → propiedades)
# ─────────────────────────────────────────────────────────────────────────────
MATERIAL_TABLE: dict = {
    189: {
        'name'        : 'HDPE (ISO 4427) SDR-17.0',
        'standard'    : 'ISO 4427',
        'schedule'    : 'SDR-17.0',
        'roughness_mm': 0.05,
    },
    360: {
        'name'        : 'Stainless Steel ASME B36.19M 10S',
        'standard'    : 'ASME B36.19M',
        'schedule'    : '10S',
        'roughness_mm': 0.04572,
    },
    77: {
        'name'        : 'Steel Sch 40 ASME B36.10M',
        'standard'    : 'ASME B36.10M',
        'schedule'    : '40',
        'roughness_mm': 0.04572,
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# Lectura y decodificación del archivo binario
# ─────────────────────────────────────────────────────────────────────────────

def read_pipe_file(path: str) -> str:
    with open(path, 'rb') as f:
        raw = f.read()
    return re.sub(r'(.)\x00', r'\1', raw.decode('latin-1', errors='replace'))

def build_line_dict(clean: str) -> dict:
    lines = {}
    for line in clean.split('\n'):
        m = re.match(r'^(\d+) (.+)', line)
        if m:
            lines[int(m.group(1))] = m.group(2)
    return lines

# ─────────────────────────────────────────────────────────────────────────────
# Tabla de especificaciones (OD / WT por spec_id y tamaño nominal)
# ─────────────────────────────────────────────────────────────────────────────

def build_spec_table(clean: str) -> dict:
    entry_re = re.compile(r'\d+ [\d.]+ in \d+ (\d+) mm ([\d.e+\-]+) ([\d.e+\-]+)')
    specs = {}
    def _parse(obj_id, start, end=None):
        s = clean.find(start)
        if s < 0: return
        e = clean.find(end, s) if end else len(clean)
        for mm, id_m, wt_m in entry_re.findall(clean[s:e]):
            try: specs[(obj_id, mm+' mm')] = (float(id_m), float(wt_m))
            except ValueError: pass
    _parse(189, 'HDPE (ISO 4427)',             'ISO 4427-1:2019')
    _parse(360, 'Stainless Steel ASME B36.19M','ASME B36.19M-2004')
    _parse(77,  'Steel Sched 40',               'esi::pipeflo::document::design_limits::velocity_limit')
    return specs

def lookup_od_wt(spec_id, nom_str, spec_table):
    if (spec_id, nom_str) in spec_table:
        id_m, wt_m = spec_table[(spec_id, nom_str)]
        return round(id_m*1e3, 2), round(wt_m*1e3, 2), round((id_m+2*wt_m)*1e3, 2)
    return None, None, None

# ─────────────────────────────────────────────────────────────────────────────
# Extracción del fluido activo
# ─────────────────────────────────────────────────────────────────────────────

def extract_fluid(clean: str) -> dict:
    """
    Extrae el fluido activo del proyecto.
    Retorna dict con nombre, temperatura, modelo y propiedades físicas.
    """
    fluid = {
        'name'                  : None,
        'temperature_c'         : None,
        'reference_pressure_bar': 0.0,
        'model'                 : None,
        'source_file'           : None,
        # Propiedades físicas estándar del fluido activo
        'density_kg_m3'         : None,
        'viscosity_cP'          : None,
        'viscosity_Pa_s'        : None,
    }

    # Nombre del fluido (Water 25 C, Air 60 F, etc.)
    fluid_name_m = re.search(
        r'(\d+) (Water \d+ [CF]|Air \d+ [CF]|Methane \d+ [CF]|[\w\s]+) '
        r'\d+ \d+ esi::pipeflo::document::fluid',
        clean
    )
    if fluid_name_m:
        fluid['name'] = fluid_name_m.group(2).strip()

    # Temperatura y archivo fuente
    temp_m = re.search(
        r'0\.0+e\+00 3 bar 1 ([\d.e+\-]+) 7 celsius 9 ([\w.]+)', clean
    )
    if temp_m:
        fluid['temperature_c']  = round(float(temp_m.group(1)), 2)
        fluid['source_file']    = temp_m.group(2)
        fluid['model']          = 'NIST fluid model'

    # Propiedades físicas según fluido y temperatura (valores NIST/estándar)
    name  = (fluid['name'] or '').lower()
    temp  = fluid['temperature_c'] or 25.0
    if 'water' in name:
        # Agua: propiedades interpoladas en función de temperatura (°C)
        # Fuente: NIST, válido 0–100°C
        rho  = 999.842 - 0.0624*temp - 0.00363*temp**2
        mu   = 1.7914e-3 * math.exp(-0.02615*(temp - 0.0))   # aprox Andrade
        # Valores conocidos exactos
        if abs(temp - 25) < 0.5:
            rho, mu = 997.05, 8.909e-4     # kg/m³, Pa·s a 25°C
        elif abs(temp - 20) < 0.5:
            rho, mu = 998.20, 1.002e-3
        elif abs(temp - 15) < 0.5:
            rho, mu = 999.10, 1.139e-3
        elif abs(temp - 10) < 0.5:
            rho, mu = 999.70, 1.308e-3
        fluid['density_kg_m3']  = round(rho, 3)
        fluid['viscosity_Pa_s'] = round(mu, 6)
        fluid['viscosity_cP']   = round(mu * 1000, 4)
    elif 'air' in name:
        # Aire: gas ideal a presión atm
        rho = 101325 / (287.058 * (temp + 273.15))
        fluid['density_kg_m3']  = round(rho, 4)
        fluid['viscosity_cP']   = round(1.716e-5 * ((temp+273.15)/273.15)**1.5 *
                                        (273.15+110.4)/(temp+273.15+110.4) * 1000, 5)
        fluid['viscosity_Pa_s'] = round(fluid['viscosity_cP'] / 1000, 8)

    return fluid

# ─────────────────────────────────────────────────────────────────────────────
# Posiciones de nodos (coordenadas de diagrama)
# ─────────────────────────────────────────────────────────────────────────────

def extract_node_positions(clean: str) -> dict:
    coord_map = {}
    for m in re.finditer(
        r'^\d+ 0\s+\d+ (Node \d+) \d+\s*\n\d+ 73 .+?6 meters 273\s*\n\d+ ([-\d.e+]+) ([-\d.e+]+)',
        clean, re.MULTILINE|re.DOTALL):
        coord_map[(round(float(m.group(2)),1), round(float(m.group(3)),1))] = m.group(1)
    n1 = re.search(
        r'0\s+0 1 \d+ Node 1 \d+\s*\n\d+ 73 .+?6 meters 273[^\n]*\n\d+ ([-\d.e+]+) ([-\d.e+]+)',
        clean, re.DOTALL)
    if n1:
        coord_map[(round(float(n1.group(1)),1), round(float(n1.group(2)),1))] = 'Node 1'
    for name, pat in [
        ('Fixed dP Device 1',   r'Fixed dP Device \d+ 0 0 154.+?BoxComp ([-\d.e+]+) ([-\d.e+]+)'),
        ('Control Valve 1',     r'GenericCont ([-\d.e+]+) ([-\d.e+]+)'),
        ('Pressure Boundary 1', r'DemandNE ([-\d.e+]+) ([-\d.e+]+)'),
        ('Centrifugal Pump 1',  r'NormalPump ([-\d.e+]+) ([-\d.e+]+)'),
    ]:
        m = re.search(pat, clean, re.DOTALL)
        if m: coord_map[(round(float(m.group(1)),1), round(float(m.group(2)),1))] = name
    tank_m = re.search(
        r'1 0 0 0 0 1 0 0 0 0 2 0 ([-\d.e+]+) ([-\d.e+]+) [-\d.e+]+ [-\d.e+]+ 0 1 196', clean)
    if tank_m:
        coord_map.setdefault(
            (round(float(tank_m.group(1)),1), round(float(tank_m.group(2)),1)), 'Tank 1')
    return coord_map

def extract_node_grid_positions(clean: str) -> dict:
    return {name: pos for pos, name in extract_node_positions(clean).items()}

# ─────────────────────────────────────────────────────────────────────────────
# Elevaciones de nodos
# ─────────────────────────────────────────────────────────────────────────────

def extract_node_elevations(clean: str) -> dict:
    elev = {}
    for m in re.finditer(
        r'^\d+ 0\s+\d+ (Node \d+) \d+\s*\n\d+ 73 1 1 0 0 1 ([\d.e+\-]+) 6 meters 273',
        clean, re.MULTILINE|re.DOTALL):
        elev[m.group(1)] = round(float(m.group(2)),4)
    n1 = re.search(
        r'0\s+0 1 \d+ Node 1 \d+\s*\n\d+ 73 1 1 0 0 1 ([\d.e+\-]+) 6 meters 273',
        clean, re.DOTALL)
    if n1: elev['Node 1'] = round(float(n1.group(1)),4)
    for name, pat in [
        ('Fixed dP Device 1',   r'Fixed dP Device \d+ 0 0 154\s*\n\d+ 73 1 1 0 0 1 ([\d.e+\-]+) 6 meters'),
        ('Control Valve 1',     r'Control Valve \d+ 154\s*\n\d+ 73 1 1 0 0 1 ([\d.e+\-]+) 6 meters'),
        ('Pressure Boundary 1', r'Pressure Boundary \d+ 154\s*\n\d+ 73 1 1 0 0 1 ([\d.e+\-]+) 6 meters'),
        ('Centrifugal Pump 1',  r'Centrifugal Pump \d+ 0 0 154\s*\n\d+ 73 1 1 0 0 1 ([\d.e+\-]+) 6 meters'),
        ('Tank 1',              r'Tank \d+ 0 0 154.*?\n\d+ 73 1 0 0 1 0 0 0 0 0 0 1 ([\d.e+\-]+)'),
    ]:
        m = re.search(pat, clean, re.DOTALL)
        if m: elev[name] = round(float(m.group(1)),4)
    return elev

# ─────────────────────────────────────────────────────────────────────────────
# Conectividad corregida (mapeo directo verificado)
# ─────────────────────────────────────────────────────────────────────────────

_COORD_LINE_TO_PIPE = {
    553:'Pipe 1',  556:'Pipe 3',  574:'Pipe 11', 600:'Pipe 12',
    608:'Pipe 13', 614:'Pipe 4',  624:'Pipe 5',  634:'Pipe 14',
    640:'Pipe 15', 646:'Pipe 6',  656:'Pipe 16', 662:'Pipe 7',
    672:'Pipe 17', 678:'Pipe 8',  688:'Pipe 18', 694:'Pipe 19',
    704:'Pipe 20', 716:'Pipe 21', 722:'Pipe 22', 734:'Pipe 23',
    744:'Pipe 24', 750:'Pipe 25', 760:'Pipe 26', 770:'Pipe 27',
    780:'Pipe 28', 794:'Pipe 29',
}

def _extract_coord_lines_raw(lines, coord_map):
    pats = [
        re.compile(r'^1 0 0 [01] \d+ \d+ 2 0 ([-\d.e+]+) ([-\d.e+]+) ([-\d.e+]+) ([-\d.e+]+) 196'),
        re.compile(r'^1 0 0 [01] \d+ \d+ 3 0 ([-\d.e+]+) ([-\d.e+]+) [-\d.e+]+ [-\d.e+]+ ([-\d.e+]+) ([-\d.e+]+) 196'),
        re.compile(r'^1 0 0 0 0 1 0 0 0 0 2 0 ([-\d.e+]+) ([-\d.e+]+) ([-\d.e+]+) ([-\d.e+]+) 0 1 196'),
    ]
    results = []
    for ln, content in sorted(lines.items()):
        for pat in pats:
            m = pat.match(content)
            if m:
                x1,y1 = round(float(m.group(1)),1), round(float(m.group(2)),1)
                x2,y2 = round(float(m.group(3)),1), round(float(m.group(4)),1)
                results.append((ln, coord_map.get((x1,y1), f'({x1},{y1})'),
                                    coord_map.get((x2,y2), f'({x2},{y2})')))
                break
    return results

def build_pipe_connectivity(lines, coord_map):
    conn = {}
    for ln, fn, tn in _extract_coord_lines_raw(lines, coord_map):
        if ln in _COORD_LINE_TO_PIPE:
            conn[_COORD_LINE_TO_PIPE[ln]] = (fn, tn)
    # Fallback Pipe 29
    if 'Pipe 29' not in conn:
        ep = _find_exit_coords(lines)
        if ep:
            nb  = _nearest_boundary(ep, coord_map)
            fn2 = _nearest_node(ep, coord_map, exclude=nb)
            conn['Pipe 29'] = (fn2, nb or 'N/D')
    return conn

def _find_exit_coords(lines):
    slns = sorted(lines.keys())
    eln  = next((ln for ln in slns if re.search(r'\bExit\b', lines[ln])), None)
    if eln is None: return None
    for ln2 in slns:
        if ln2 <= eln: continue
        ec = re.search(r'1ExitOpaque ([-\d.e+]+) ([-\d.e+]+)', lines[ln2])
        if ec: return (round(float(ec.group(1)),1), round(float(ec.group(2)),1))
    return None

def _nearest_boundary(pos, coord_map):
    best, bd = None, float('inf')
    for (cx,cy), name in coord_map.items():
        if 'Pressure Boundary' in name or 'Tank' in name:
            d = abs(cx-pos[0])+abs(cy-pos[1])
            if d < bd: bd, best = d, name
    return best

def _nearest_node(pos, coord_map, exclude=None):
    cands = [(abs(cx-pos[0])+abs(cy-pos[1]), name)
             for (cx,cy), name in coord_map.items()
             if name != exclude and abs(cx-pos[0])+abs(cy-pos[1]) < 1.5]
    cands.sort()
    return cands[0][1] if cands else 'N/D'

# ─────────────────────────────────────────────────────────────────────────────
# Fittings y K por cañería
# ─────────────────────────────────────────────────────────────────────────────

def extract_fittings_for_pipe(name_line: int, next_pipe_line, lines: dict) -> list:
    OBJ_FITTING = {
        '212': {'category':'Bend','name':'Mitre Bend @ 90°','k_value':60.0,'k_type':'LeD'},
        '562': {'category':'Bend','name':'Mitre Bend @ 90°','k_value':60.0,'k_type':'LeD'},
        '566': {'category':'Bend','name':'Mitre Bend @ 45°','k_value':15.0,'k_type':'LeD'},
        '710': {'category':'Bend','name':'Mitre Bend @ 60°','k_value':25.0,'k_type':'LeD'},
        '728': {'category':'Bend','name':'Mitre Bend @ 90°','k_value':60.0,'k_type':'LeD'},
    }
    slns     = sorted(lines.keys())
    fittings = []
    red_re   = re.compile(r'^(?:\d+ )?(Fitting) \d+ (Reducer - (?:Contraction|Enlargement)) \d+ 1 ([\d.e+\-]+) \d+ mm')
    led_re   = re.compile(r'^(?:\d+ )?(Valve|Bend|Check Valve|Fitting|Other) \d+ (.+?) \d+ ([\d.]+e[+\-]\d+|\d+\.\d+)')
    ball_re  = re.compile(r'^1 0\s+124 \d+')
    obj_re   = re.compile(r'^1 0\s+109 (\d+) \d+')
    geom_re  = re.compile(r'^1 0\s+(118|119) \d+ 1 ([\d.e+\-]+) \d+ mm')
    bfly_re  = re.compile(r'^\d+ \d+ 1ButterflyBlack ')

    for ln in slns:
        if ln <= name_line: continue
        if next_pipe_line and ln >= next_pipe_line: break
        c = lines[ln]
        m = red_re.match(c)
        if m:
            try:
                diam = round(float(m.group(3)), 0)
                fittings.append({'category':'Fitting',
                                 'name': f"{m.group(2).strip()} ({diam:.0f} mm)",
                                 'k_value':None,'k_type':'geometry'})
            except: pass
            continue
        m = led_re.match(c)
        if m:
            cat, name = m.group(1), m.group(2).strip()
            if any(x in name for x in ('Reducer','Contraction','Enlargement')):
                fittings.append({'category':'Fitting','name':name,'k_value':None,'k_type':'geometry'})
            else:
                try:
                    raw    = float(m.group(3))
                    k_type = 'K' if (cat=='Fitting' and any(x in name for x in ('Exit','Entrance'))) else 'LeD'
                    fittings.append({'category':cat,'name':name,
                                     'k_value':round(raw,4) if k_type=='K' else round(raw,2),
                                     'k_type':k_type})
                except: pass
            continue
        if ball_re.match(c):
            fittings.append({'category':'Valve','name':'Ball','k_value':3.0,'k_type':'LeD'})
            continue
        m = obj_re.match(c)
        if m:
            obj_id = m.group(1)
            if obj_id in OBJ_FITTING:
                fittings.append(dict(OBJ_FITTING[obj_id]))
            continue
        m = geom_re.match(c)
        if m:
            try:
                diam  = round(float(m.group(2)), 0)
                rtype = 'Reducer - Contraction' if m.group(1)=='118' else 'Reducer - Enlargement'
                fittings.append({'category':'Fitting','name':f"{rtype} ({diam:.0f} mm)",
                                 'k_value':None,'k_type':'geometry'})
            except: pass
            continue
        if bfly_re.match(c):
            fittings.append({'category':'Valve','name':'Butterfly','k_value':45.0,'k_type':'LeD'})

    return fittings

def _f_turb(roughness_mm: float, id_mm: float) -> float:
    if id_mm <= 0: return 0.0112
    rel = roughness_mm / id_mm
    return (-2*math.log10(rel/3.7))**-2 if rel > 0 else 0.0112

def compute_k_total(fittings, id_mm=None, roughness_mm=0.05) -> float:
    if not fittings: return 0.0
    f_T = _f_turb(roughness_mm, id_mm) if id_mm else 0.0112
    k   = sum((f['k_value'] if f['k_type']=='K' else f_T*f['k_value'])
              for f in fittings if f['k_value'] is not None)
    return round(k, 4)

def summarise_fittings(fittings) -> str:
    parts = []
    for f in fittings:
        if f['k_value'] is None:         parts.append(f['name'])
        elif f['k_type'] == 'K':         parts.append(f"{f['name']} (K={f['k_value']})")
        else:                             parts.append(f"{f['name']} (Le/D={f['k_value']})")
    return '; '.join(parts) if parts else '—'

# ─────────────────────────────────────────────────────────────────────────────
# Propiedades de cañerías (con material y rugosidad)
# ─────────────────────────────────────────────────────────────────────────────

def extract_pipe_properties(lines: dict, spec_table: dict) -> list:
    clean_text   = '\n'.join(f'{k} {v}' for k,v in sorted(lines.items()))
    pipe_re      = re.compile(r'^(\d+) (?:\d+ )?(Pipe \d+) 171$', re.MULTILINE)
    slns         = sorted(lines.keys())
    all_name_lns = sorted(int(m.group(1)) for m in pipe_re.finditer(clean_text))
    pipes = []

    for m in pipe_re.finditer(clean_text):
        name_line = int(m.group(1))
        pname     = m.group(2)
        idx          = all_name_lns.index(name_line)
        next_pipe_ln = all_name_lns[idx+1] if idx+1 < len(all_name_lns) else None

        diam_str, length_m, spec_id = 'N/D', None, None
        for ln in slns:
            if ln <= name_line: continue
            if next_pipe_ln and ln >= next_pipe_ln: break
            c = lines[ln]
            if not re.match(r'^73 1 1 0 0 \d+', c): continue
            dm = re.search(r'(\d+) mm', c)
            if dm: diam_str = dm.group(1)+' mm'
            sm = re.search(r'\b77 (\d{2,})\b', c)
            if sm: spec_id = int(sm.group(1))
            lm = re.search(r'\b1 ([\d.e+\-]+) (?:\d+ \d+ )?6 meters', c)
            if lm: length_m = round(float(lm.group(1)),4)
            break

        id_mm, wt_mm, od_mm = lookup_od_wt(spec_id, diam_str, spec_table)
        material  = MATERIAL_TABLE.get(spec_id, {})
        roughness = material.get('roughness_mm', 0.05)

        fittings         = extract_fittings_for_pipe(name_line, next_pipe_ln, lines)
        fittings_summary = summarise_fittings(fittings)
        k_computed       = compute_k_total(fittings, id_mm, roughness)
        k_total          = USER_K_OVERRIDES.get(pname, k_computed)

        pipes.append({
            'name'              : pname,
            'name_line'         : name_line,
            'diameter'          : diam_str,
            'od_mm'             : od_mm,
            'wt_mm'             : wt_mm,
            'id_mm'             : id_mm,
            'length_m'          : length_m,
            'material_name'     : material.get('name', 'N/D'),
            'material_standard' : material.get('standard', 'N/D'),
            'material_schedule' : material.get('schedule', 'N/D'),
            'roughness_mm'      : roughness,
            'spec_obj_id'       : spec_id,
            'fittings'          : fittings,
            'fittings_summary'  : fittings_summary,
            'k_total'           : k_total,
        })
    return pipes

# ─────────────────────────────────────────────────────────────────────────────
# Componentes especiales
# ─────────────────────────────────────────────────────────────────────────────

def extract_special_components(clean: str, lines: dict) -> dict:
    result = {'tanks':[],'pumps':[],'fixed_dp':[],'control_valves':[],'pressure_boundaries':[]}
    # Tank
    tp = clean.find('Tank 1 0 0 154 1 0')
    if tp >= 0:
        tb  = clean[tp:tp+400]
        em  = re.search(r'73 1 0 0 1 0 0 0 0 0 0 1 ([\d.e+\-]+) 0 0 6 meters', tb)
        pm  = re.search(r'73 1 0 0 1 0 0 0 0 1 ([\d.e+\-]+) 163', tb)
        lm  = re.search(r'73 1 0 0 1 0 0 0 0 0 0 1 ([\d.e+\-]+) 0 0 6 meters 0 0 171', tb)
        result['tanks'].append({'name':'Tank 1',
            'elevation_m':round(float(em.group(1)),4) if em else None,
            'surface_pressure_kpa_abs':round(float(pm.group(1)),2) if pm else None,
            'pressure_unit':'kPa (abs)',
            'liquid_level_m':round(float(lm.group(1)),4) if lm else None})
    # Pump
    pp = clean.find('Centrifugal Pump 1 0 0 154')
    if pp >= 0:
        pb  = clean[pp:pp+500]
        sm  = re.search(r'73 1 1 0 0 1 ([\d.e+\-]+) 6 meters 0 0 154', pb)
        dm  = re.search(r'73 1 1 0 0 1 ([\d.e+\-]+) 6 meters 0 0 0 0 240', pb)
        fm  = re.search(r'1 ([\d.e+\-]+) 0 0 0 0 250', pb)
        result['pumps'].append({'name':'Centrifugal Pump 1',
            'operation_mode':'flow' if 'operation_mode_flow' in pb else 'unknown',
            'flow_rate':round(float(fm.group(1)),2) if fm else None,
            'flow_rate_unit':'m3/h',
            'suction_elevation_m':round(float(sm.group(1)),4) if sm else None,
            'discharge_elevation_m':round(float(dm.group(1)),4) if dm else None})
    # Fixed dP
    fp = clean.find('Fixed dP Device 1 0 0 154')
    if fp >= 0:
        fb    = clean[fp:fp+300]
        elevs = re.findall(r'73 1 1 0 0 1 ([\d.e+\-]+) 6 meters', fb)
        dpm   = re.search(r'73 1 0 0 1 0 0 0 0 0 0 1 ([\d.e+\-]+) 0 0 0 (\d+) (bar|kPa|Pa|psi)', fb)
        result['fixed_dp'].append({'name':'Fixed dP Device 1',
            'inlet_elevation_m':round(float(elevs[0]),4) if elevs else None,
            'outlet_elevation_m':round(float(elevs[1]),4) if len(elevs)>1 else None,
            'pressure_drop':round(float(dpm.group(1)),4) if dpm else None,
            'pressure_drop_unit':dpm.group(3) if dpm else 'bar'})
    # Control Valve
    cp = clean.find('Control Valve 1 154')
    if cp >= 0:
        cb   = clean[cp:cp+1200]
        em   = re.search(r'73 1 1 0 0 1 ([\d.e+\-]+) 6 meters', cb)
        mode = 'Fixed Cv' if 'operation_mode_fixed_cv' in cb else 'unknown'
        fc_m = re.search(r'([\d.]+e[+\-]\d+) 2 (Cv|Kv|Cc)', cb) or \
               re.search(r'(\d+\.\d+) 2 (Cv|Kv|Cc)', cb)
        result['control_valves'].append({'name':'Control Valve 1',
            'elevation_m':round(float(em.group(1)),4) if em else None,
            'operation_mode':mode,
            'flow_coefficient':round(float(fc_m.group(1)),2) if fc_m else None,
            'flow_coefficient_unit':fc_m.group(2) if fc_m else '—'})
    # Pressure Boundary
    bp = clean.find('Pressure Boundary 1 154')
    if bp >= 0:
        bb  = clean[bp:bp+300]
        em  = re.search(r'73 1 1 0 0 1 ([\d.e+\-]+) 6 meters', bb)
        pm  = re.search(r'1 ([\d.e+\-]+) 163', bb)
        result['pressure_boundaries'].append({'name':'Pressure Boundary 1',
            'elevation_m':round(float(em.group(1)),4) if em else None,
            'pressure_kpa_abs':round(float(pm.group(1)),2) if pm else None,
            'pressure_unit':'kPa (abs)',
            'operation_mode':'pressure'})
    return result

# ─────────────────────────────────────────────────────────────────────────────
# Función principal de extracción
# ─────────────────────────────────────────────────────────────────────────────

def extract_all(filepath: str) -> dict:
    """Retorna {'pipes':[], 'nodes':[], 'fluid':{}, tanks, pumps, ...}"""
    clean        = read_pipe_file(filepath)
    lines        = build_line_dict(clean)
    spec_table   = build_spec_table(clean)
    coord_map    = extract_node_positions(clean)
    elevations   = extract_node_elevations(clean)
    grid_pos     = extract_node_grid_positions(clean)
    connectivity = build_pipe_connectivity(lines, coord_map)
    pipes_props  = extract_pipe_properties(lines, spec_table)
    special      = extract_special_components(clean, lines)
    fluid        = extract_fluid(clean)

    pipes_full = []
    for pipe in sorted(pipes_props, key=lambda p: p['name_line']):
        pname    = pipe['name']
        fn, tn   = connectivity.get(pname, ('N/D','N/D'))
        k_total  = USER_K_OVERRIDES.get(pname, pipe['k_total'])
        pipes_full.append({
            'name'              : pname,
            'diameter'          : pipe['diameter'],
            'od_mm'             : pipe['od_mm'],
            'wt_mm'             : pipe['wt_mm'],
            'id_mm'             : pipe['id_mm'],
            'length_m'          : pipe['length_m'],
            'material_name'     : pipe['material_name'],
            'material_standard' : pipe['material_standard'],
            'material_schedule' : pipe['material_schedule'],
            'roughness_mm'      : pipe['roughness_mm'],
            'from'              : fn,
            'to'                : tn,
            'fittings'          : pipe['fittings'],
            'fittings_summary'  : pipe['fittings_summary'],
            'k_total'           : k_total,
        })

    all_names = set(elevations.keys()) | set(grid_pos.keys())
    nodes = [{'name': n,
              'elevation_m': elevations.get(n),
              'grid_x': grid_pos.get(n, (None,None))[0],
              'grid_y': grid_pos.get(n, (None,None))[1]}
             for n in sorted(all_names)]

    return {'pipes': pipes_full, 'nodes': nodes, 'fluid': fluid, **special}

# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN DE EXPORTACIÓN  →  todas las variables listas para uso externo
# ─────────────────────────────────────────────────────────────────────────────

def export_variables(filepath: str) -> dict:
    """
    Lee un archivo .pipe de PipeFlo y retorna UN DICCIONARIO PLANO con
    todas las variables del proyecto, listo para usarse en cualquier
    programa externo (cálculo hidráulico, pandas, MATLAB, etc.).

    Estructura retornada:
    ──────────────────────────────────────────────────────
    {
      # ── FLUIDO ──────────────────────────────────────
      'fluid_name'           : str      # 'Water 25 C'
      'fluid_temperature_c'  : float    # 25.0
      'fluid_density_kg_m3'  : float    # 997.05
      'fluid_viscosity_cP'   : float    # 0.8909
      'fluid_viscosity_Pa_s' : float    # 0.0008909

      # ── CAÑERÍAS (lista de dicts) ────────────────────
      'pipes': [
        {
          'name'        : 'Pipe 1'
          'diameter'    : '160 mm'
          'od_mm'       : 160.0
          'wt_mm'       : 9.5
          'id_mm'       : 141.0
          'id_m'        : 0.141
          'area_m2'     : 0.015621      # sección transversal interior
          'length_m'    : 1.404
          'roughness_mm': 0.05
          'roughness_m' : 0.00005
          'material'    : 'HDPE...'
          'from'        : 'Tank 1'
          'to'          : 'Centrifugal Pump 1'
          'k_total'     : 1.028
          'fittings'    : [...]         # lista detallada
        }, ...
      ]

      # ── NODOS (lista de dicts) ───────────────────────
      'nodes': [
        { 'name':'Node 1', 'elevation_m':4.724, 'grid_x':-1.0, 'grid_y':4.0 }, ...
      ]

      # ── ACCESO RÁPIDO POR NOMBRE ─────────────────────
      'pipe_by_name'  : { 'Pipe 1': {...}, ... }
      'node_by_name'  : { 'Node 1': {...}, ... }

      # ── COMPONENTES ESPECIALES ───────────────────────
      'tanks'              : [...]
      'pumps'              : [...]
      'fixed_dp'           : [...]
      'control_valves'     : [...]
      'pressure_boundaries': [...]

      # ── MATERIALES ÚNICOS ────────────────────────────
      'materials': [
        { 'name':'HDPE...', 'standard':'ISO 4427', 'schedule':'SDR-17.0',
          'roughness_mm':0.05, 'pipes': ['Pipe 1', ...] }, ...
      ]
    }
    """
    import math as _math

    data = extract_all(filepath)
    fluid = data.get('fluid', {})

    # ── Enriquecer cada pipe con variables derivadas ───────────────────────
    enriched_pipes = []
    for p in data['pipes']:
        ep = dict(p)
        # Diámetro interior en metros
        id_m    = (p['id_mm'] / 1000) if p['id_mm'] else None
        area_m2 = (_math.pi * id_m**2 / 4) if id_m else None
        ep['id_m']      = round(id_m, 6)    if id_m    else None
        ep['area_m2']   = round(area_m2, 8) if area_m2 else None
        ep['roughness_m'] = round(p['roughness_mm'] / 1000, 8)
        ep['material']  = p['material_name']
        enriched_pipes.append(ep)

    # ── Índices rápidos ───────────────────────────────────────────────────
    pipe_by_name = {p['name']: p for p in enriched_pipes}
    node_by_name = {n['name']: n for n in data['nodes']}

    # ── Tabla de materiales únicos ─────────────────────────────────────────
    mat_index = {}
    for p in enriched_pipes:
        mname = p.get('material_name', 'N/D')
        if mname not in mat_index:
            mat_index[mname] = {
                'name'        : mname,
                'standard'    : p.get('material_standard','N/D'),
                'schedule'    : p.get('material_schedule','N/D'),
                'roughness_mm': p.get('roughness_mm', None),
                'pipes'       : [],
            }
        mat_index[mname]['pipes'].append(p['name'])

    return {
        # Fluido
        'fluid_name'           : fluid.get('name'),
        'fluid_temperature_c'  : fluid.get('temperature_c'),
        'fluid_density_kg_m3'  : fluid.get('density_kg_m3'),
        'fluid_viscosity_cP'   : fluid.get('viscosity_cP'),
        'fluid_viscosity_Pa_s' : fluid.get('viscosity_Pa_s'),
        'fluid_model'          : fluid.get('model'),
        'fluid_source'         : fluid.get('source_file'),
        # Cañerías
        'pipes'                : enriched_pipes,
        'pipe_by_name'         : pipe_by_name,
        # Nodos
        'nodes'                : data['nodes'],
        'node_by_name'         : node_by_name,
        # Componentes especiales
        'tanks'                : data.get('tanks', []),
        'pumps'                : data.get('pumps', []),
        'fixed_dp'             : data.get('fixed_dp', []),
        'control_valves'       : data.get('control_valves', []),
        'pressure_boundaries'  : data.get('pressure_boundaries', []),
        # Materiales
        'materials'            : list(mat_index.values()),
    }

def extract_pipes(filepath: str) -> list:
    return extract_all(filepath)['pipes']

# ─────────────────────────────────────────────────────────────────────────────
# Salidas: pantalla, CSV, Excel, JSON
# ─────────────────────────────────────────────────────────────────────────────

def print_table(data: dict):
    pipes = data['pipes']
    nodes = data['nodes']
    fluid = data.get('fluid', {})
    sep   = '─' * 148

    # Fluido
    print(); print(sep)
    print(f"{'FLUIDO ACTIVO':^148}"); print(sep)
    print(f"  Nombre     : {fluid.get('name','N/D')}")
    print(f"  Temperatura: {fluid.get('temperature_c','N/D')} °C")
    print(f"  Densidad   : {fluid.get('density_kg_m3','N/D')} kg/m³")
    print(f"  Viscosidad : {fluid.get('viscosity_cP','N/D')} cP  ({fluid.get('viscosity_Pa_s','N/D')} Pa·s)")
    print(f"  Modelo     : {fluid.get('model','N/D')} ({fluid.get('source_file','')})"); print(sep); print()

    # Cañerías
    print(sep); print(f"{'CAÑERÍAS':^148}"); print(sep)
    print(f"{'Cañería':<10} {'D.Nom.':<9} {'OD(mm)':<8} {'WT(mm)':<8} {'ID(mm)':<8} "
          f"{'Long.(m)':<11} {'K_total':<8} {'Material':<32} "
          f"{'Nodo Inicio':<22} {'Nodo Fin':<22} Fittings")
    print(sep)
    for p in pipes:
        od  = f"{p['od_mm']:.2f}"    if p['od_mm']    else 'N/D'
        wt  = f"{p['wt_mm']:.2f}"    if p['wt_mm']    else 'N/D'
        id_ = f"{p['id_mm']:.2f}"    if p['id_mm']    else 'N/D'
        lg  = f"{p['length_m']:.4f}" if p['length_m'] else 'N/D'
        kt  = f"{p['k_total']:.4f}"  if p['k_total']  else '0.0000'
        mat = p.get('material_name','N/D')[:30]
        print(f"{p['name']:<10} {p['diameter']:<9} {od:<8} {wt:<8} {id_:<8} "
              f"{lg:<11} {kt:<8} {mat:<32} "
              f"{str(p['from']):<22} {str(p['to']):<22} {p['fittings_summary']}")
    print(sep); print(f"  Total: {len(pipes)} cañerías\n")

    # Nodos
    print(sep); print(f"{'NODOS / ELEVACIONES / POSICIÓN EN GRILLA':^148}"); print(sep)
    print(f"  {'Nodo':<32} {'Elevación (m)':<18} {'Grid X':<10} Grid Y"); print(sep)
    for n in nodes:
        elev = f"{n['elevation_m']:.4f}" if n['elevation_m'] is not None else 'N/D'
        gx   = str(n['grid_x']) if n['grid_x'] is not None else 'N/D'
        gy   = str(n['grid_y']) if n['grid_y'] is not None else 'N/D'
        print(f"  {n['name']:<32} {elev:<18} {gx:<10} {gy}")
    print(sep); print(f"  Total: {len(nodes)} nodos\n")

    for title, key, fields in [
        ('ESTANQUES',          'tanks',
         [('Elevación','elevation_m','m'),
          ('Presión superficial','surface_pressure_kpa_abs','pressure_unit'),
          ('Nivel de líquido','liquid_level_m','m')]),
        ('BOMBAS CENTRÍFUGAS', 'pumps',
         [('Modo operación','operation_mode','—'),
          ('Caudal','flow_rate','flow_rate_unit'),
          ('Elev. succión','suction_elevation_m','m'),
          ('Elev. descarga','discharge_elevation_m','m')]),
        ('FIXED dP DEVICES',   'fixed_dp',
         [('Elev. entrada','inlet_elevation_m','m'),
          ('Elev. salida','outlet_elevation_m','m'),
          ('Caída de presión','pressure_drop','pressure_drop_unit')]),
        ('VÁLVULAS DE CONTROL','control_valves',
         [('Elevación','elevation_m','m'),
          ('Modo','operation_mode','—'),
          ('Coef. de flujo','flow_coefficient','flow_coefficient_unit')]),
        ('PRESSURE BOUNDARIES','pressure_boundaries',
         [('Elevación','elevation_m','m'),
          ('Presión','pressure_kpa_abs','pressure_unit'),
          ('Modo','operation_mode','—')]),
    ]:
        items = data.get(key, [])
        if not items: continue
        print(sep); print(f"{title:^148}"); print(sep)
        for item in items:
            print(f"  {item['name']}")
            for label, field, uf in fields:
                val  = item.get(field,'—')
                unit = item.get(uf,'') if uf not in ('m','—') else uf
                if unit == '—': unit = ''
                print(f"    {label:<30}: {val} {unit}")
        print(sep); print()


def save_csv(data: dict, path: str):
    base  = path.rsplit('.',1)[0] if '.' in os.path.basename(path) else path
    pipes = data['pipes']
    nodes = data['nodes']
    fluid = data.get('fluid', {})

    with open(base+'_canerias.csv','w',newline='',encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['Cañería','D. Nominal','OD (mm)','WT (mm)','ID (mm)',
                    'Longitud (m)','K Total','Material','Rugosidad (mm)',
                    'Nodo Inicio','Nodo Fin','Fittings / Válvulas'])
        for p in pipes:
            w.writerow([p['name'],p['diameter'],p['od_mm'] or '',p['wt_mm'] or '',
                        p['id_mm'] or '',p['length_m'] or '',p['k_total'] or '',
                        p.get('material_name',''),p.get('roughness_mm',''),
                        p['from'],p['to'],p['fittings_summary']])

    with open(base+'_nodos.csv','w',newline='',encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['Nodo','Elevación (m)','Grid X','Grid Y'])
        for n in nodes:
            w.writerow([n['name'],n['elevation_m'] or '',n['grid_x'] or '',n['grid_y'] or ''])

    with open(base+'_componentes.csv','w',newline='',encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['Tipo','Nombre','Parámetro','Valor','Unidad'])
        # Fluido
        for k, label, unit in [
            ('name','Nombre',''), ('temperature_c','Temperatura','°C'),
            ('density_kg_m3','Densidad','kg/m³'), ('viscosity_cP','Viscosidad','cP'),
        ]:
            w.writerow(['Fluido', fluid.get('name','—'), label, fluid.get(k,'—'), unit])
        # Componentes
        for t in data.get('tanks',[]):
            w.writerow(['Tank',t['name'],'Elevación',t['elevation_m'],'m'])
            w.writerow(['Tank',t['name'],'Presión superficial',t['surface_pressure_kpa_abs'],t['pressure_unit']])
            w.writerow(['Tank',t['name'],'Nivel de líquido',t['liquid_level_m'],'m'])
        for p in data.get('pumps',[]):
            w.writerow(['Pump',p['name'],'Modo',p['operation_mode'],'—'])
            w.writerow(['Pump',p['name'],'Caudal',p['flow_rate'],p['flow_rate_unit']])
            w.writerow(['Pump',p['name'],'Elev. succión',p['suction_elevation_m'],'m'])
            w.writerow(['Pump',p['name'],'Elev. descarga',p['discharge_elevation_m'],'m'])
        for d in data.get('fixed_dp',[]):
            w.writerow(['Fixed dP',d['name'],'Elev. entrada',d['inlet_elevation_m'],'m'])
            w.writerow(['Fixed dP',d['name'],'Elev. salida',d['outlet_elevation_m'],'m'])
            w.writerow(['Fixed dP',d['name'],'ΔP',d['pressure_drop'],d['pressure_drop_unit']])
        for v in data.get('control_valves',[]):
            w.writerow(['CV',v['name'],'Elevación',v['elevation_m'],'m'])
            w.writerow(['CV',v['name'],'Modo',v['operation_mode'],'—'])
            w.writerow(['CV',v['name'],'Coef. flujo',v['flow_coefficient'],v['flow_coefficient_unit']])
        for pb in data.get('pressure_boundaries',[]):
            w.writerow(['PB',pb['name'],'Elevación',pb['elevation_m'],'m'])
            w.writerow(['PB',pb['name'],'Presión',pb['pressure_kpa_abs'],pb['pressure_unit']])

    print(f"CSV: {base}_canerias.csv | {base}_nodos.csv | {base}_componentes.csv")


def save_excel(data: dict, path: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("pip install openpyxl"); return

    pipes = data['pipes']
    nodes = data['nodes']
    fluid = data.get('fluid', {})
    wb    = openpyxl.Workbook()

    h_fill = PatternFill(start_color='1F4E79',end_color='1F4E79',fill_type='solid')
    h_font = Font(color='FFFFFF',bold=True,size=11)
    a_fill = PatternFill(start_color='EBF3FB',end_color='EBF3FB',fill_type='solid')
    thin   = Side(style='thin')
    brd    = Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr    = Alignment(horizontal='center',vertical='center',wrap_text=True)
    lft    = Alignment(horizontal='left',  vertical='center',wrap_text=True)

    def _hdr(ws, headers, widths):
        for col,(h,w) in enumerate(zip(headers,widths),1):
            c=ws.cell(row=1,column=col,value=h)
            c.font,c.fill,c.alignment,c.border=h_font,h_fill,ctr,brd
            ws.column_dimensions[c.column_letter].width=w
        ws.row_dimensions[1].height=22; ws.freeze_panes='A2'

    def _row(ws, i, vals, lcols=()):
        fill=a_fill if i%2==0 else None
        for col,val in enumerate(vals,1):
            c=ws.cell(row=i,column=col,value=val)
            c.border=brd; c.alignment=lft if col in lcols else ctr
            if fill: c.fill=fill

    # Cañerías
    ws1=wb.active; ws1.title='Cañerías'
    _hdr(ws1,['Cañería','D. Nominal','OD (mm)','WT (mm)','ID (mm)','Longitud (m)',
               'K Total','Material','Rugosidad (mm)','Nodo Inicio','Nodo Fin','Fittings'],
         [12,12,9,9,9,13,9,34,14,24,24,65])
    for i,p in enumerate(pipes,2):
        _row(ws1,i,[p['name'],p['diameter'],p['od_mm'],p['wt_mm'],p['id_mm'],
                    p['length_m'],p['k_total'],p.get('material_name'),p.get('roughness_mm'),
                    p['from'],p['to'],p['fittings_summary']],lcols=(12,))
    ws1.cell(row=len(pipes)+2,column=1,value=f'Total: {len(pipes)} cañerías').font=Font(bold=True)

    # Nodos
    ws2=wb.create_sheet('Nodos')
    _hdr(ws2,['Nodo','Elevación (m)','Grid X','Grid Y'],[32,16,10,10])
    for i,n in enumerate(nodes,2):
        _row(ws2,i,[n['name'],n['elevation_m'],n['grid_x'],n['grid_y']])
    ws2.cell(row=len(nodes)+2,column=1,value=f'Total: {len(nodes)} nodos').font=Font(bold=True)

    # Componentes + Fluido
    ws3=wb.create_sheet('Componentes y Fluido')
    _hdr(ws3,['Tipo','Nombre','Parámetro','Valor','Unidad'],[20,24,30,18,12])
    row=2
    # Fluido
    for label,key,unit in [('Nombre','name',''),('Temperatura','temperature_c','°C'),
                            ('Densidad','density_kg_m3','kg/m³'),
                            ('Viscosidad','viscosity_cP','cP'),
                            ('Viscosidad','viscosity_Pa_s','Pa·s'),
                            ('Modelo','model',''),('Fuente','source_file','')]:
        _row(ws3,row,['Fluido',fluid.get('name','—'),label,fluid.get(key,'—'),unit]); row+=1
    row+=1
    for tipo, items, fn in [
        ('Tank', data.get('tanks',[]),
         lambda t:[('Elevación',t['elevation_m'],'m'),
                   ('Presión superficial',t['surface_pressure_kpa_abs'],t['pressure_unit']),
                   ('Nivel de líquido',t['liquid_level_m'],'m')]),
        ('Bomba',data.get('pumps',[]),
         lambda p:[('Modo',p['operation_mode'],'—'),
                   ('Caudal',p['flow_rate'],p['flow_rate_unit']),
                   ('Elev. succión',p['suction_elevation_m'],'m'),
                   ('Elev. descarga',p['discharge_elevation_m'],'m')]),
        ('Fixed dP',data.get('fixed_dp',[]),
         lambda d:[('Elev. entrada',d['inlet_elevation_m'],'m'),
                   ('Elev. salida',d['outlet_elevation_m'],'m'),
                   ('ΔP',d['pressure_drop'],d['pressure_drop_unit'])]),
        ('C. Valve',data.get('control_valves',[]),
         lambda v:[('Elevación',v['elevation_m'],'m'),
                   ('Modo',v['operation_mode'],'—'),
                   ('Coef. flujo',v['flow_coefficient'],v['flow_coefficient_unit'])]),
        ('PB',data.get('pressure_boundaries',[]),
         lambda pb:[('Elevación',pb['elevation_m'],'m'),
                    ('Presión',pb['pressure_kpa_abs'],pb['pressure_unit'])]),
    ]:
        for item in items:
            for param,val,unit in fn(item):
                _row(ws3,row,[tipo,item['name'],param,val,unit]); row+=1
        if items: row+=1

    wb.save(path)
    print(f"Excel guardado: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI interactivo
# ─────────────────────────────────────────────────────────────────────────────

def seleccionar_archivo_pipe():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    archivos   = [f for f in os.listdir(script_dir) if f.endswith('.pipe')]
    if not archivos:
        print("No se encontraron archivos .pipe en el directorio del script.")
        sys.exit(1)
    print("\nArchivos disponibles:")
    for i,a in enumerate(archivos,1):
        print(f"  [{i}] {a}")
    while True:
        try:
            s = int(input("\nSelecciona un archivo: "))
            if 1 <= s <= len(archivos):
                return os.path.join(script_dir, archivos[s-1])
        except ValueError: pass
        print("Selección inválida.")

def seleccionar_formato_salida():
    print("\n¿Deseas guardar el resultado?")
    print("  [1] CSV  (tres archivos)")
    print("  [2] Excel  (.xlsx con cuatro hojas)")
    print("  [3] JSON  (todas las variables)")
    print("  [4] No guardar")
    while True:
        op = input("Selecciona: ").strip()
        if op in ('1','2','3','4'):
            return {'1':'csv','2':'excel','3':'json','4':None}[op]
        print("Opción inválida.")

def main():
    archivo = seleccionar_archivo_pipe()
    print(f"\nProcesando: {archivo}")
    data = extract_all(archivo)
    print_table(data)
    fmt  = seleccionar_formato_salida()
    if fmt == 'csv':
        nombre = input("Nombre base (ej: resultado): ").strip()
        save_csv(data, nombre)
    elif fmt == 'excel':
        nombre = input("Nombre Excel (ej: resultado.xlsx): ").strip()
        if not nombre.endswith('.xlsx'): nombre += '.xlsx'
        save_excel(data, nombre)
    elif fmt == 'json':
        nombre = input("Nombre JSON (ej: resultado.json): ").strip()
        if not nombre.endswith('.json'): nombre += '.json'
        variables = export_variables(archivo)
        # Serializar fittings (lista de dicts)
        def _serial(obj):
            if isinstance(obj, float) and (obj != obj): return None
            return obj
        with open(nombre,'w',encoding='utf-8') as f:
            json.dump(variables, f, ensure_ascii=False, indent=2, default=_serial)
        print(f"JSON guardado: {nombre}")
    return data

if __name__ == '__main__':
    main()